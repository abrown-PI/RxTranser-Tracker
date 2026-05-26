#!/usr/bin/env python
"""
Import patient transfer data from 5 Excel workbooks and output data.js
"""

import json
import re
import os
import sys
from datetime import datetime, date

import openpyxl

# ── File definitions ────────────────────────────────────────────────────────
DOWNLOADS = r"C:\Users\ABrown\Downloads"

FILES = [
    {
        "path": os.path.join(DOWNLOADS, "Weekly shipments to Florida (2).xlsx"),
        "origin": "Seminole",
        "type": "florida",
    },
    {
        "path": os.path.join(DOWNLOADS, "Tucson Transfer Patient Status (1).xlsx"),
        "origin": "Tucson",
        "type": "tucson",
    },
    {
        "path": os.path.join(DOWNLOADS, "Virginia Beach Transfer Patient Status (1).xlsx"),
        "origin": "Virginia Beach",
        "type": "virginia",
    },
    {
        "path": os.path.join(DOWNLOADS, "Lancaster Patient Status (1).xlsx"),
        "origin": "Lancaster",
        "type": "lancaster",
    },
    {
        "path": os.path.join(DOWNLOADS, "Jamestown Transfer Patient Status.xlsx"),
        "origin": "Jamestown",
        "type": "jamestown",
    },
]

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "data.js")


# ── Helpers ─────────────────────────────────────────────────────────────────

def parse_date(val):
    """Return 'YYYY-MM-DD' string or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    # Try common formats
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y", "%m-%d-%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # If it's not a date-like string (e.g. "biller"), return None
    return None


def clean_name(val):
    """Clean and title-case a patient name. Returns empty string if blank."""
    if val is None:
        return ""
    s = str(val).strip()
    if not s:
        return ""
    # Remove known suffixes/tags
    for tag in ["(nrx-estradiol)", "(nrx-progest)", "(nrx)", "2rxs", "(NRX)", "(Nrx)"]:
        s = s.replace(tag, "")
    # Remove anything in parens that looks like a tag
    s = re.sub(r"\(nrx[^)]*\)", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\b2rxs\b", "", s, flags=re.IGNORECASE)
    s = s.strip()
    if not s:
        return ""
    # Ensure space after comma  "Ciraci,Toni" → "Ciraci, Toni"
    s = re.sub(r",\s*", ", ", s)
    # Collapse multiple spaces
    s = re.sub(r"\s{2,}", " ", s).strip()
    # Title-case each part
    parts = s.split(", ")
    parts = [p.strip().title() for p in parts]
    return ", ".join(parts)


def normalize_ship_to(val, origin):
    """Normalize ship-to value."""
    if val is None:
        return "Pharmacy"
    s = str(val).strip().lower()
    if not s:
        return "Pharmacy"
    if "patient" in s:
        return "Patient"
    # Location names mean shipping to the pharmacy at that location
    for loc in ["florida", "arizona", "virginia", "lancaster", "jamestown",
                "seminole", "tucson", "ny", "new york"]:
        if loc in s:
            return "Pharmacy"
    if s in ("pharmacy",):
        return "Pharmacy"
    # Default: keep original but capitalize
    return "Pharmacy"


def normalize_sema(val):
    if val is None:
        return "NA"
    s = str(val).strip().lower()
    if s in ("yes", "y"):
        return "Yes"
    if s in ("no", "n"):
        return "No"
    if s in ("", "na", "n/a"):
        return "NA"
    return "NA"


def normalize_cold(val):
    if val is None:
        return False
    s = str(val).strip().lower()
    return s in ("yes", "y", "true", "1")


def normalize_paid(val):
    if val is None:
        return "Pending"
    s = str(val).strip()
    if not s:
        return "Pending"
    if s.lower() in ("x",):
        return "Pending"
    return s


def safe_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() == "none":
        return ""
    return s


def derive_status(date_rcvd, date_shipped, date_rcvd_erie):
    if date_rcvd:
        return "Delivered"
    if date_shipped:
        return "Shipped"
    if date_rcvd_erie:
        return "In Progress"
    return "New"


def derive_created_at(tx_email, date_rcvd_erie, antic_ship):
    for d in (tx_email, date_rcvd_erie, antic_ship):
        if d:
            return d + "T00:00:00"
    return "2026-01-01T00:00:00"


def find_header_row(rows_iter, name_col_candidates):
    """Read rows until we find one that looks like a header (contains a patient name column)."""
    for row in rows_iter:
        vals = [str(v).strip().lower() if v else "" for v in row]
        for candidate in name_col_candidates:
            if candidate.lower() in vals:
                return row
    return None


def build_col_map(header_row):
    """Build a dict of lowered-header → column-index."""
    m = {}
    for i, v in enumerate(header_row):
        if v is not None:
            key = str(v).strip().lower()
            if key not in m:
                m[key] = i
            else:
                # Duplicate header name – store with suffix
                m[key + f"__{i}"] = i
    return m


# ── Per-file extraction ────────────────────────────────────────────────────

def get_col(row, col_map, *candidates, default=None):
    """Get value from row by trying multiple column name candidates."""
    for c in candidates:
        cl = c.lower()
        if cl in col_map:
            idx = col_map[cl]
            if idx < len(row):
                return row[idx]
        # Try partial match
        for key, idx in col_map.items():
            if cl in key and idx < len(row):
                return row[idx]
    return default


def extract_florida(wb, origin):
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        col_map = build_col_map(header)

        for row in rows[1:]:
            if row is None:
                continue
            name_raw = get_col(row, col_map, "patient name")
            name = clean_name(name_raw)
            if not name:
                continue
            records.append({
                "patientName": name,
                "originLocation": origin,
                "fillLocation": "Erie",
                "shipTo": normalize_ship_to(get_col(row, col_map, "ship to"), origin),
                "semaDrAuthorized": normalize_sema(get_col(row, col_map, "if sema, dr. authorized", "if sema dr. authorized")),
                "txEmailedDate": parse_date(get_col(row, col_map, "tx emailed to erie")),
                "dateReceivedAtErie": parse_date(get_col(row, col_map, "date received")),
                "patientStatus": safe_str(get_col(row, col_map, "patient status")) or "Pending",
                "coldItem": normalize_cold(get_col(row, col_map, "cold item")),
                "anticipatedShipDate": parse_date(get_col(row, col_map, "anticipated ship date")),
                "dateShipped": parse_date(get_col(row, col_map, "date shipped")),
                "trackingNumber": safe_str(get_col(row, col_map, "tracking number")),
                "paid": normalize_paid(get_col(row, col_map, "paid")),
                "dateReceivedByPatient": parse_date(get_col(row, col_map, "date rcv'd in fl")),
                "drug": "" if safe_str(get_col(row, col_map, "drug")).lower() in ("x", "xx", "") else safe_str(get_col(row, col_map, "drug")),
                "qty": "",
                "sheetName": sheet_name,
            })
    return records


def extract_tucson(wb, origin):
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        col_map = build_col_map(header)

        for row in rows[1:]:
            if row is None:
                continue
            name_raw = get_col(row, col_map, "patient name")
            name = clean_name(name_raw)
            if not name:
                continue
            # Drug: combine Medication + Medication (if Custom) or use Drug col
            medication = safe_str(get_col(row, col_map, "medication"))
            med_custom = safe_str(get_col(row, col_map, "medication (if custom)"))
            drug_col = safe_str(get_col(row, col_map, "drug"))
            # drug_col from "Drug" header is usually just "x" marker
            if drug_col.lower() in ("x", "xx"):
                drug_col = ""
            drug = med_custom if med_custom else (medication if medication else drug_col)

            records.append({
                "patientName": name,
                "originLocation": origin,
                "fillLocation": "Erie",
                "shipTo": normalize_ship_to(get_col(row, col_map, "ship to"), origin),
                "semaDrAuthorized": normalize_sema(get_col(row, col_map, "if sema, dr. authorized", "if sema dr. authorized")),
                "txEmailedDate": parse_date(get_col(row, col_map, "transfer sent to erie")),
                "dateReceivedAtErie": parse_date(get_col(row, col_map, "date/time order received")),
                "patientStatus": safe_str(get_col(row, col_map, "patient status")) or "Pending",
                "coldItem": normalize_cold(get_col(row, col_map, "cold item")),
                "anticipatedShipDate": parse_date(get_col(row, col_map, "anticipated ship date")),
                "dateShipped": parse_date(get_col(row, col_map, "date shipped")),
                "trackingNumber": safe_str(get_col(row, col_map, "tracking number")),
                "paid": normalize_paid(get_col(row, col_map, "paid")),
                "dateReceivedByPatient": parse_date(get_col(row, col_map, "date received in az")),
                "drug": drug,
                "qty": safe_str(get_col(row, col_map, "quantity/packaging")),
                "sheetName": sheet_name,
            })
    return records


def extract_virginia(wb, origin):
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        if header is None or all(v is None for v in header):
            continue
        col_map = build_col_map(header)

        # Virginia Beach has a second "Drug" column (index 5) with actual drug info
        # The first "Drug" col (index 1) is just "x" markers
        # Find the drug column that's NOT just a marker
        drug_col_idx = None
        for key, idx in col_map.items():
            if "drug" in key:
                # Check if this is the later drug column (higher index, after patient name)
                patient_idx = col_map.get("patient name", 3)
                if idx > patient_idx:
                    drug_col_idx = idx
                    break
        # If no second drug column found, try col index 5 directly from data
        if drug_col_idx is None:
            # Check if col 5 header is "Drug"
            if len(header) > 5 and header[5] is not None and "drug" in str(header[5]).lower():
                drug_col_idx = 5

        for row in rows[1:]:
            if row is None:
                continue
            name_raw = get_col(row, col_map, "patient name")
            name = clean_name(name_raw)
            if not name:
                continue

            drug_val = ""
            if drug_col_idx is not None and drug_col_idx < len(row):
                drug_val = safe_str(row[drug_col_idx])
            # If drug_val is just "x", clear it
            if drug_val.lower() in ("x", "xx"):
                drug_val = ""

            records.append({
                "patientName": name,
                "originLocation": origin,
                "fillLocation": "Erie",
                "shipTo": normalize_ship_to(get_col(row, col_map, "ship to"), origin),
                "semaDrAuthorized": normalize_sema(get_col(row, col_map, "if sema, dr. authorized", "if sema dr. authorized")),
                "txEmailedDate": None,  # VB doesn't have this column
                "dateReceivedAtErie": parse_date(get_col(row, col_map, "date/time order received")),
                "patientStatus": safe_str(get_col(row, col_map, "patient status")) or "Pending",
                "coldItem": normalize_cold(get_col(row, col_map, "cold item")),
                "anticipatedShipDate": parse_date(get_col(row, col_map, "anticipated ship date")),
                "dateShipped": parse_date(get_col(row, col_map, "date shipped")),
                "trackingNumber": safe_str(get_col(row, col_map, "tracking number")),
                "paid": normalize_paid(get_col(row, col_map, "paid")),
                "dateReceivedByPatient": parse_date(get_col(row, col_map, "date received in va")),
                "drug": drug_val,
                "qty": "",
                "sheetName": sheet_name,
            })
    return records


def extract_lancaster(wb, origin):
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        if header is None or all(v is None for v in header):
            continue
        col_map = build_col_map(header)

        # Lancaster has two "Drug" columns: col 0 (marker "x") and col 10 (actual drug info)
        # Find the drug column with actual drug info (after Patient Name)
        drug_col_idx = None
        patient_idx = None
        for key, idx in col_map.items():
            if "patient name" in key:
                patient_idx = idx
        for key, idx in col_map.items():
            if "drug" in key and patient_idx is not None and idx > patient_idx:
                drug_col_idx = idx
                break

        for row in rows[1:]:
            if row is None:
                continue
            name_raw = get_col(row, col_map, "patient name (last, first)", "patient name")
            name = clean_name(name_raw)
            if not name:
                continue

            drug_val = ""
            if drug_col_idx is not None and drug_col_idx < len(row):
                drug_val = safe_str(row[drug_col_idx])
            if drug_val.lower() in ("x", "xx"):
                drug_val = ""

            # Payment method column
            paid_val = get_col(row, col_map, "payment method")
            if paid_val is None:
                paid_val = get_col(row, col_map, "paid")

            records.append({
                "patientName": name,
                "originLocation": origin,
                "fillLocation": "Erie",
                "shipTo": normalize_ship_to(get_col(row, col_map, "ship to"), origin),
                "semaDrAuthorized": normalize_sema(get_col(row, col_map, "if sema, dr. authorized", "if sema dr. authorized")),
                "txEmailedDate": parse_date(get_col(row, col_map, "tx sent to erie")),
                "dateReceivedAtErie": parse_date(get_col(row, col_map, "date order rcv'd")),
                "patientStatus": safe_str(get_col(row, col_map, "patient status")) or "Pending",
                "coldItem": normalize_cold(get_col(row, col_map, "cold item")),
                "anticipatedShipDate": parse_date(get_col(row, col_map, "anticipated ship date")),
                "dateShipped": parse_date(get_col(row, col_map, "date shipped")),
                "trackingNumber": safe_str(get_col(row, col_map, "tracking number")),
                "paid": normalize_paid(paid_val),
                "dateReceivedByPatient": parse_date(get_col(row, col_map, "date rcv'd in lancaster")),
                "drug": drug_val,
                "qty": "",
                "sheetName": sheet_name,
            })
    return records


def extract_jamestown(wb, origin):
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        if header is None or all(v is None for v in header):
            continue
        col_map = build_col_map(header)

        for row in rows[1:]:
            if row is None:
                continue
            name_raw = get_col(row, col_map, "patient name")
            name = clean_name(name_raw)
            if not name:
                continue

            # Drug column - skip "x" marker values
            drug_val = safe_str(get_col(row, col_map, "drug"))
            if drug_val.lower() in ("x", "xx"):
                drug_val = ""

            # Paid column - Jamestown has two "Paid" columns (index 0 marker, index 13 actual)
            # Try to get the second one
            paid_val = None
            paid_indices = [idx for key, idx in col_map.items() if "paid" in key]
            if len(paid_indices) > 1:
                # Use the later paid column
                later_idx = max(paid_indices)
                if later_idx < len(row):
                    paid_val = row[later_idx]
            else:
                paid_val = get_col(row, col_map, "paid")
            # If paid is just "x"/"X", treat as pending
            if paid_val is not None and str(paid_val).strip().lower() == "x":
                paid_val = None

            records.append({
                "patientName": name,
                "originLocation": origin,
                "fillLocation": "Erie",
                "shipTo": normalize_ship_to(get_col(row, col_map, "ship to"), origin),
                "semaDrAuthorized": normalize_sema(get_col(row, col_map, "if sema, dr. authorized", "if sema dr. authorized")),
                "txEmailedDate": parse_date(get_col(row, col_map, "tx sent to erie")),
                "dateReceivedAtErie": parse_date(get_col(row, col_map, "date order rcv'd")),
                "patientStatus": safe_str(get_col(row, col_map, "patient status")) or "Pending",
                "coldItem": normalize_cold(get_col(row, col_map, "cold item")),
                "anticipatedShipDate": parse_date(get_col(row, col_map, "anticipated ship date")),
                "dateShipped": parse_date(get_col(row, col_map, "date shipped")),
                "trackingNumber": safe_str(get_col(row, col_map, "tracking number")),
                "paid": normalize_paid(paid_val),
                "dateReceivedByPatient": parse_date(get_col(row, col_map, "date rcv'd in ny")),
                "drug": drug_val,
                "qty": "",
                "sheetName": sheet_name,
            })
    return records


# ── Main ────────────────────────────────────────────────────────────────────

EXTRACTORS = {
    "florida": extract_florida,
    "tucson": extract_tucson,
    "virginia": extract_virginia,
    "lancaster": extract_lancaster,
    "jamestown": extract_jamestown,
}


def main():
    all_records = []

    for file_def in FILES:
        fpath = file_def["path"]
        origin = file_def["origin"]
        ftype = file_def["type"]
        print(f"Reading {origin}: {os.path.basename(fpath)} ... ", end="", flush=True)

        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        extractor = EXTRACTORS[ftype]
        records = extractor(wb, origin)
        wb.close()

        print(f"{len(records)} records")
        all_records.extend(records)

    # Assign IDs and compute derived fields
    for i, rec in enumerate(all_records, start=1):
        rec["id"] = i
        rec["status"] = derive_status(
            rec["dateReceivedByPatient"],
            rec["dateShipped"],
            rec["dateReceivedAtErie"],
        )
        rec["createdAt"] = derive_created_at(
            rec["txEmailedDate"],
            rec["dateReceivedAtErie"],
            rec["anticipatedShipDate"],
        )

    # Write data.js
    # Build JS-safe JSON with boolean lowercase
    lines = []
    lines.append("const IMPORT_DATA = [")
    for rec in all_records:
        # Order fields nicely
        obj = {
            "id": rec["id"],
            "patientName": rec["patientName"],
            "originLocation": rec["originLocation"],
            "fillLocation": rec["fillLocation"],
            "shipTo": rec["shipTo"],
            "semaDrAuthorized": rec["semaDrAuthorized"],
            "txEmailedDate": rec["txEmailedDate"],
            "dateReceivedAtErie": rec["dateReceivedAtErie"],
            "patientStatus": rec["patientStatus"],
            "coldItem": rec["coldItem"],
            "anticipatedShipDate": rec["anticipatedShipDate"],
            "dateShipped": rec["dateShipped"],
            "trackingNumber": rec["trackingNumber"],
            "paid": rec["paid"],
            "dateReceivedByPatient": rec["dateReceivedByPatient"],
            "drug": rec["drug"],
            "qty": rec["qty"],
            "status": rec["status"],
            "createdAt": rec["createdAt"],
            "sheetName": rec["sheetName"],
        }
        # Convert to JS-compatible JSON (booleans lowercase, null not None)
        js_str = json.dumps(obj, ensure_ascii=False)
        # json.dumps already outputs true/false/null correctly for JS
        lines.append(f"  {js_str},")
    lines.append("];")

    js_content = "\n".join(lines) + "\n"

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(js_content)

    print(f"\nTotal records: {len(all_records)}")
    print(f"Output written to: {OUTPUT_FILE}")

    # Print breakdown by origin
    from collections import Counter
    origin_counts = Counter(r["originLocation"] for r in all_records)
    for origin, count in sorted(origin_counts.items()):
        print(f"  {origin}: {count}")

    # Print status breakdown
    status_counts = Counter(r["status"] for r in all_records)
    for status, count in sorted(status_counts.items()):
        print(f"  {status}: {count}")


if __name__ == "__main__":
    main()
